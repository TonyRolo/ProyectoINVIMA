import openpyxl

book = openpyxl.load_workbook('prueba.xlsx')
sheet = book.active

dataExcel = []

def dataDownload(dataExcel):
    for row in range(1, sheet.max_row):
        _row = [row,]
        for col in sheet.iter_cols(2, 3):
            _row.append(col[row].value)
        dataExcel.append(_row)
    return dataExcel

def excelUpdate(dataExcel):
    for i in range(0, len(dataExcel)):
        if(array[i][2] is None):
            sheet[f'C{i}'].value = array[i][2]
            print(dataExcel[i][2])
    book.save('prueba.xlsx')

dataDownload(dataExcel)
excelUpdate(dataExcel) 